import csv
import uuid
from io import BytesIO, StringIO
from pathlib import Path
from collections import OrderedDict

from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from django.utils.decorators import method_decorator
from openpyxl import Workbook, load_workbook

from LMS.roles import get_logged_in_name, get_student_profile, get_teacher_profile, is_admin_user
from LMS.upload_utils import validate_uploaded_file
from LMS.views import access_denied_response
from .models import EnrollmentYear, Faculty, Student, Subject
from .forms import FirstLoginPasswordChangeForm


class AdminOnlyMixin:
    def dispatch(self, request, *args, **kwargs):
        if not is_admin_user(request.user):
            return access_denied_response(request, 'Only administrators can access this page.')
        return super().dispatch(request, *args, **kwargs)


class StudentListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = 'student/student_list.html'
    context_object_name = 'students'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user
        queryset = Student.objects.select_related('faculty', 'enrollment_batch').order_by(
            'faculty__name',
            'enrollment_batch__year',
            'name',
        )
        search = (self.request.GET.get('q') or '').strip()
        faculty_id = (self.request.GET.get('faculty') or '').strip()
        enrollment_id = (self.request.GET.get('year') or '').strip()

        if search:
            queryset = queryset.filter(name__icontains=search)
        if faculty_id:
            queryset = queryset.filter(faculty_id=faculty_id)
        if enrollment_id:
            queryset = queryset.filter(enrollment_batch_id=enrollment_id)

        if is_admin_user(user):
            return queryset
        teacher = get_teacher_profile(user)
        if teacher:
            return queryset.filter(faculty__in=teacher.faculties.all()).distinct()
        return queryset.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        view_mode = self.request.GET.get('view', 'grouped')
        if view_mode not in {'grouped', 'flat'}:
            view_mode = 'grouped'

        context['view_mode'] = view_mode

        grouped_students_map = OrderedDict()
        for student in context['page_obj']:
            faculty_name = student.faculty.name if student.faculty else 'No Faculty'
            enrollment_year = (
                str(student.enrollment_batch.year) if student.enrollment_batch else 'No Enrollment Year'
            )
            group_key = (faculty_name, enrollment_year)
            grouped_students_map.setdefault(group_key, []).append(student)

        context['grouped_students'] = [
            {
                'faculty': faculty,
                'enrollment_year': enrollment_year,
                'students': students,
            }
            for (faculty, enrollment_year), students in grouped_students_map.items()
        ]

        query_dict = self.request.GET.copy()
        query_dict.pop('page', None)
        context['filter_query'] = query_dict.urlencode()
        context['search_query'] = (self.request.GET.get('q') or '').strip()
        context['selected_faculty'] = (self.request.GET.get('faculty') or '').strip()
        context['selected_year'] = (self.request.GET.get('year') or '').strip()
        context['faculties'] = Faculty.objects.order_by('name')
        context['enrollment_years'] = EnrollmentYear.objects.order_by('year')
        return context

class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = 'student/student_detail.html'
    context_object_name = 'student'

    def get_queryset(self):
        user = self.request.user
        queryset = Student.objects.select_related('faculty', 'enrollment_batch', 'user')
        if is_admin_user(user):
            return queryset

        student = get_student_profile(user)
        if student:
            return queryset.filter(pk=student.pk)

        teacher = get_teacher_profile(user)
        if teacher:
            return queryset.filter(faculty__in=teacher.faculties.all()).distinct()

        return queryset.none()

class StudentCreateView(LoginRequiredMixin, AdminOnlyMixin, CreateView):
    model = Student
    template_name = 'student/student_form.html'
    fields = ['name', 'roll_number', 'enrollment_batch', 'faculty', 'age', 'email', 'dp', 'address']
    success_url = reverse_lazy('student:student-create')

class StudentUpdateView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):
    model = Student
    template_name = 'student/student_form.html'
    fields = ['name', 'roll_number', 'enrollment_batch', 'faculty', 'age', 'email', 'dp', 'address']
    success_url = reverse_lazy('student:student-list')

class StudentDeleteView(LoginRequiredMixin, AdminOnlyMixin, DeleteView):
    model = Student
    template_name = 'student/student_confirm_delete.html'
    success_url = reverse_lazy('student:student-list')


class StudentImportView(LoginRequiredMixin, AdminOnlyMixin, View):
    template_name = 'student/student_import.html'
    IMPORT_SESSION_KEY = 'student_import_preview'

    def _stage_upload(self, upload):
        staging_dir = Path(settings.MEDIA_ROOT) / 'import_staging'
        staging_dir.mkdir(parents=True, exist_ok=True)
        file_name = f'student-{uuid.uuid4().hex}.xlsx'
        file_path = staging_dir / file_name
        with file_path.open('wb') as destination:
            for chunk in upload.chunks():
                destination.write(chunk)
        return str(file_path)

    def _cleanup_staged_file(self, file_path):
        if not file_path:
            return
        path_obj = Path(file_path)
        if path_obj.exists():
            path_obj.unlink()

    def _process_workbook(self, workbook, commit=False):
        worksheet = workbook.active
        created_count = 0
        updated_count = 0
        skipped_count = 0
        preview_rows = []
        error_rows = []

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value in (None, '') for value in row):
                continue

            name = str(row[0]).strip() if row[0] else ''
            roll_number = str(row[1]).strip() if row[1] else None
            enrollment_year_raw = row[2]
            faculty_name = str(row[3]).strip() if row[3] else None
            age_raw = row[4]
            email = str(row[5]).strip() if row[5] else None
            address = str(row[6]).strip() if row[6] else None

            if not name:
                skipped_count += 1
                error_rows.append({'row': row_index, 'error': 'Missing required name.'})
                continue

            enrollment_batch = None
            if enrollment_year_raw not in (None, ''):
                try:
                    year_value = int(enrollment_year_raw)
                except (TypeError, ValueError):
                    skipped_count += 1
                    error_rows.append({'row': row_index, 'error': 'Invalid enrollment year.'})
                    continue
                enrollment_batch, _ = EnrollmentYear.objects.get_or_create(year=year_value)

            faculty = None
            if faculty_name:
                faculty, _ = Faculty.objects.get_or_create_case_insensitive(faculty_name)

            age = None
            if age_raw not in (None, ''):
                try:
                    age = int(age_raw)
                except (TypeError, ValueError):
                    skipped_count += 1
                    error_rows.append({'row': row_index, 'error': 'Invalid age value.'})
                    continue

            defaults = {
                'name': name,
                'enrollment_batch': enrollment_batch,
                'faculty': faculty,
                'age': age,
                'email': email,
                'address': address,
            }

            operation = 'create'
            if roll_number and Student.objects.filter(roll_number=roll_number).exists():
                operation = 'update'
            elif not roll_number and email and Student.objects.filter(email=email).exists():
                operation = 'update'

            if len(preview_rows) < 100:
                preview_rows.append(
                    {
                        'row': row_index,
                        'name': name,
                        'roll_number': roll_number or '-',
                        'email': email or '-',
                        'operation': operation,
                    }
                )

            if not commit:
                if operation == 'create':
                    created_count += 1
                else:
                    updated_count += 1
                continue

            try:
                if roll_number:
                    _, created = Student.objects.update_or_create(
                        roll_number=roll_number,
                        defaults=defaults,
                    )
                elif email:
                    _, created = Student.objects.update_or_create(
                        email=email,
                        defaults={**defaults, 'roll_number': None},
                    )
                else:
                    Student.objects.create(
                        **defaults,
                        roll_number=None,
                    )
                    created = True
            except Exception:
                skipped_count += 1
                error_rows.append({'row': row_index, 'error': 'Database write failed for this row.'})
                continue

            if created:
                created_count += 1
            else:
                updated_count += 1

        return {
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'preview_rows': preview_rows,
            'error_rows': error_rows,
        }

    def _get_preview_context(self, request):
        return request.session.get(self.IMPORT_SESSION_KEY, {})

    def get(self, request):
        context = self._get_preview_context(request)
        return render(request, self.template_name, context)

    def post(self, request):
        action = request.POST.get('action', 'preview')

        if action == 'commit':
            preview_data = request.session.get(self.IMPORT_SESSION_KEY)
            if not preview_data or not preview_data.get('staged_file_path'):
                messages.error(request, 'No staged import found. Please preview an XLSX file first.')
                return redirect('student:student-import')

            try:
                workbook = load_workbook(preview_data['staged_file_path'], data_only=True)
            except Exception:
                self._cleanup_staged_file(preview_data.get('staged_file_path'))
                request.session.pop(self.IMPORT_SESSION_KEY, None)
                messages.error(request, 'Unable to read staged file. Please upload and preview again.')
                return redirect('student:student-import')

            result = self._process_workbook(workbook, commit=True)
            self._cleanup_staged_file(preview_data.get('staged_file_path'))
            request.session.pop(self.IMPORT_SESSION_KEY, None)

            messages.success(
                request,
                f'Import committed. Created: {result["created_count"]}, Updated: {result["updated_count"]}, Skipped: {result["skipped_count"]}.',
            )
            return redirect('student:student-list')

        upload = request.FILES.get('xlsx_file')
        try:
            validate_uploaded_file(
                upload,
                allowed_extensions={'.xlsx'},
                allowed_mime_types={
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/octet-stream',
                },
                max_size_bytes=settings.MAX_IMPORT_UPLOAD_BYTES,
            )
        except Exception as exc:
            messages.error(request, str(exc))
            return redirect('student:student-import')

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            messages.error(request, 'Unable to read the file. Please use the provided template.')
            return redirect('student:student-import')

        previous_staged_file = request.session.get(self.IMPORT_SESSION_KEY, {}).get('staged_file_path')
        self._cleanup_staged_file(previous_staged_file)
        staged_file_path = self._stage_upload(upload)
        result = self._process_workbook(workbook, commit=False)
        request.session[self.IMPORT_SESSION_KEY] = {
            'staged_file_path': staged_file_path,
            'uploaded_file_name': upload.name,
            'created_count': result['created_count'],
            'updated_count': result['updated_count'],
            'skipped_count': result['skipped_count'],
            'preview_rows': result['preview_rows'],
            'error_rows': result['error_rows'],
        }

        messages.info(request, 'Preview generated. Review changes and commit when ready.')
        return redirect('student:student-import')


class StudentImportErrorReportView(LoginRequiredMixin, AdminOnlyMixin, View):
    def get(self, request):
        preview_data = request.session.get(StudentImportView.IMPORT_SESSION_KEY)
        if not preview_data:
            messages.error(request, 'No preview is available for download.')
            return redirect('student:student-import')

        error_rows = preview_data.get('error_rows', [])
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['row', 'error'])
        for error in error_rows:
            writer.writerow([error.get('row', ''), error.get('error', '')])

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="student_import_errors.csv"'
        return response


class StudentImportTemplateView(LoginRequiredMixin, AdminOnlyMixin, View):
    def get(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Students'

        worksheet.append([
            'name',
            'roll_number',
            'enrollment_year',
            'faculty',
            'age',
            'email',
            'address',
        ])
        worksheet.append([
            'Aarav Sharma',
            'CS-2026-001',
            2026,
            'Computer Science',
            20,
            'aarav@example.com',
            'Kathmandu',
        ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        return response


class SubjectListView(LoginRequiredMixin, AdminOnlyMixin, ListView):
    model = Subject
    template_name = 'student/subject_list.html'
    context_object_name = 'subjects'
    paginate_by = 25

    def get_queryset(self):
        queryset = Subject.objects.select_related('faculty').order_by('faculty__name', 'name')
        search = (self.request.GET.get('q') or '').strip()
        faculty_id = (self.request.GET.get('faculty') or '').strip()
        if search:
            queryset = queryset.filter(name__icontains=search)
        if faculty_id:
            queryset = queryset.filter(faculty_id=faculty_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query_dict = self.request.GET.copy()
        query_dict.pop('page', None)
        context['filter_query'] = query_dict.urlencode()
        context['search_query'] = (self.request.GET.get('q') or '').strip()
        context['selected_faculty'] = (self.request.GET.get('faculty') or '').strip()
        context['faculties'] = Faculty.objects.order_by('name')
        return context


class SubjectImportView(LoginRequiredMixin, AdminOnlyMixin, View):
    template_name = 'student/subject_import.html'
    IMPORT_SESSION_KEY = 'subject_import_preview'

    def _stage_upload(self, upload):
        staging_dir = Path(settings.MEDIA_ROOT) / 'import_staging'
        staging_dir.mkdir(parents=True, exist_ok=True)
        file_name = f'subject-{uuid.uuid4().hex}.xlsx'
        file_path = staging_dir / file_name
        with file_path.open('wb') as destination:
            for chunk in upload.chunks():
                destination.write(chunk)
        return str(file_path)

    def _cleanup_staged_file(self, file_path):
        if not file_path:
            return
        path_obj = Path(file_path)
        if path_obj.exists():
            path_obj.unlink()

    def _process_workbook(self, workbook, commit=False):
        worksheet = workbook.active
        created_count = 0
        updated_count = 0
        skipped_count = 0
        preview_rows = []
        error_rows = []

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value in (None, '') for value in row):
                continue

            name = str(row[0]).strip() if row[0] else ''
            abbreviation = str(row[1]).strip() if row[1] else None
            code = str(row[2]).strip() if row[2] else None
            faculty_name = str(row[3]).strip() if row[3] else None

            if not name:
                skipped_count += 1
                error_rows.append({'row': row_index, 'error': 'Missing required subject name.'})
                continue

            faculty = None
            if faculty_name:
                faculty, _ = Faculty.objects.get_or_create_case_insensitive(faculty_name)

            defaults = {
                'name': name,
                'abbreviation': abbreviation,
                'code': code,
                'faculty': faculty,
            }

            if code:
                lookup = {'code': code}
                operation = 'update' if Subject.objects.filter(code=code).exists() else 'create'
            elif abbreviation:
                lookup = {'abbreviation': abbreviation}
                operation = 'update' if Subject.objects.filter(abbreviation=abbreviation).exists() else 'create'
            else:
                lookup = {'name': name}
                operation = 'update' if Subject.objects.filter(name=name).exists() else 'create'

            if len(preview_rows) < 100:
                preview_rows.append(
                    {
                        'row': row_index,
                        'name': name,
                        'abbreviation': abbreviation or '-',
                        'code': code or '-',
                        'operation': operation,
                    }
                )

            if not commit:
                if operation == 'create':
                    created_count += 1
                else:
                    updated_count += 1
                continue

            try:
                _, created = Subject.objects.update_or_create(defaults=defaults, **lookup)
            except Exception:
                skipped_count += 1
                error_rows.append({'row': row_index, 'error': 'Database write failed for this row.'})
                continue

            if created:
                created_count += 1
            else:
                updated_count += 1

        return {
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'preview_rows': preview_rows,
            'error_rows': error_rows,
        }

    def _get_preview_context(self, request):
        return request.session.get(self.IMPORT_SESSION_KEY, {})

    def get(self, request):
        context = self._get_preview_context(request)
        return render(request, self.template_name, context)

    def post(self, request):
        action = request.POST.get('action', 'preview')

        if action == 'commit':
            preview_data = request.session.get(self.IMPORT_SESSION_KEY)
            if not preview_data or not preview_data.get('staged_file_path'):
                messages.error(request, 'No staged import found. Please preview an XLSX file first.')
                return redirect('student:subject-import')

            try:
                workbook = load_workbook(preview_data['staged_file_path'], data_only=True)
            except Exception:
                self._cleanup_staged_file(preview_data.get('staged_file_path'))
                request.session.pop(self.IMPORT_SESSION_KEY, None)
                messages.error(request, 'Unable to read staged file. Please upload and preview again.')
                return redirect('student:subject-import')

            result = self._process_workbook(workbook, commit=True)
            self._cleanup_staged_file(preview_data.get('staged_file_path'))
            request.session.pop(self.IMPORT_SESSION_KEY, None)

            messages.success(
                request,
                f'Import committed. Created: {result["created_count"]}, Updated: {result["updated_count"]}, Skipped: {result["skipped_count"]}.',
            )
            return redirect('student:subject-list')

        upload = request.FILES.get('xlsx_file')
        try:
            validate_uploaded_file(
                upload,
                allowed_extensions={'.xlsx'},
                allowed_mime_types={
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/octet-stream',
                },
                max_size_bytes=settings.MAX_IMPORT_UPLOAD_BYTES,
            )
        except Exception as exc:
            messages.error(request, str(exc))
            return redirect('student:subject-import')

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            messages.error(request, 'Unable to read the file. Please use the provided template.')
            return redirect('student:subject-import')

        previous_staged_file = request.session.get(self.IMPORT_SESSION_KEY, {}).get('staged_file_path')
        self._cleanup_staged_file(previous_staged_file)
        staged_file_path = self._stage_upload(upload)
        result = self._process_workbook(workbook, commit=False)
        request.session[self.IMPORT_SESSION_KEY] = {
            'staged_file_path': staged_file_path,
            'uploaded_file_name': upload.name,
            'created_count': result['created_count'],
            'updated_count': result['updated_count'],
            'skipped_count': result['skipped_count'],
            'preview_rows': result['preview_rows'],
            'error_rows': result['error_rows'],
        }

        messages.info(request, 'Preview generated. Review changes and commit when ready.')
        return redirect('student:subject-import')


class SubjectImportErrorReportView(LoginRequiredMixin, AdminOnlyMixin, View):
    def get(self, request):
        preview_data = request.session.get(SubjectImportView.IMPORT_SESSION_KEY)
        if not preview_data:
            messages.error(request, 'No preview is available for download.')
            return redirect('student:subject-import')

        error_rows = preview_data.get('error_rows', [])
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['row', 'error'])
        for error in error_rows:
            writer.writerow([error.get('row', ''), error.get('error', '')])

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="subject_import_errors.csv"'
        return response


class SubjectImportTemplateView(LoginRequiredMixin, AdminOnlyMixin, View):
    def get(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Subjects'

        worksheet.append(['name', 'abbreviation', 'code', 'faculty'])
        worksheet.append(['Discrete Mathematics', 'DM', 'MATH101', 'Computer Science'])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="subject_import_template.xlsx"'
        return response


@method_decorator(login_required, name='dispatch')
class FirstLoginPasswordChangeView(View):
    """View for changing password on first login"""
    template_name = 'student/first_login_password_change.html'
    
    def get(self, request):
        """Display password change form"""
        try:
            student = request.user.student_profile
            if not student.is_first_login:
                messages.info(request, 'You have already changed your password.')
                return redirect('student:student-list')
        except AttributeError:
            # User is not a student, redirect to home
            messages.error(request, 'Only students can access this page.')
            return redirect('/')
        
        form = FirstLoginPasswordChangeForm(request.user)
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        """Process password change"""
        try:
            student = request.user.student_profile
        except AttributeError:
            messages.error(request, 'Only students can access this page.')
            return redirect('/')
        
        form = FirstLoginPasswordChangeForm(request.user, request.POST)
        
        if form.is_valid():
            form.save()
            
            # Mark first login as False
            student.is_first_login = False
            student.save()
            
            messages.success(request, 'Your password has been changed successfully. Please login again.')
            return redirect('login')
        
        return render(request, self.template_name, {'form': form})

