from django.contrib import admin
from django.urls import reverse, path
from django.utils.html import format_html, mark_safe
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from io import BytesIO
from openpyxl import Workbook, load_workbook
from .models import EnrollmentYear, Faculty, Student, Subject
from .admin_actions import (
    import_students_from_csv, 
    generate_credentials_pdf, 
    generate_csv_template
)


class CSVImportForm:
    """Simple form handler for CSV import"""
    pass


@admin.register(Faculty)
class FacultyAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)


@admin.register(EnrollmentYear)
class EnrollmentYearAdmin(admin.ModelAdmin):
    list_display = ('year',)
    search_fields = ('year',)


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ('name', 'abbreviation', 'code', 'faculty')
    search_fields = ('name', 'abbreviation', 'code', 'faculty__name')
    list_filter = ('faculty',)
    change_list_template = 'admin/subject/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-xlsx/',
                self.admin_site.admin_view(self.import_xlsx),
                name='student_subject_import_xlsx',
            ),
            path(
                'download-xlsx-template/',
                self.admin_site.admin_view(self.download_xlsx_template),
                name='student_subject_download_xlsx_template',
            ),
        ]
        return custom_urls + urls

    def import_xlsx(self, request):
        if request.method == 'POST':
            upload = request.FILES.get('xlsx_file')
            if not upload:
                messages.error(request, 'Please select an XLSX file to import.')
                return render(request, 'admin/subject_import.html', {'admin_site': self.admin_site})

            if not upload.name.lower().endswith('.xlsx'):
                messages.error(request, 'Only .xlsx files are supported.')
                return render(request, 'admin/subject_import.html', {'admin_site': self.admin_site})

            try:
                workbook = load_workbook(upload, data_only=True)
            except Exception:
                messages.error(request, 'Unable to read the file. Please use the provided template.')
                return render(request, 'admin/subject_import.html', {'admin_site': self.admin_site})

            worksheet = workbook.active
            created_count = 0
            updated_count = 0
            skipped_count = 0

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row or all(value in (None, '') for value in row):
                    continue

                name = str(row[0]).strip() if row[0] else ''
                abbreviation = str(row[1]).strip() if row[1] else None
                code = str(row[2]).strip() if row[2] else None
                faculty_name = str(row[3]).strip() if row[3] else None

                if not name:
                    skipped_count += 1
                    continue

                faculty = None
                if faculty_name:
                    faculty, _ = Faculty.objects.get_or_create_case_insensitive(faculty_name)

                defaults = {
                    'name': name,
                    'abbreviation': abbreviation,
                    'code': code,
                    'faculty': faculty,
                }

                lookup = None
                if code:
                    lookup = {'code': code}
                elif abbreviation:
                    lookup = {'abbreviation': abbreviation}
                else:
                    lookup = {'name': name}

                try:
                    _, created = Subject.objects.update_or_create(defaults=defaults, **lookup)
                except Exception:
                    skipped_count += 1
                    continue

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            messages.success(
                request,
                f'Import finished. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}.',
            )
            return redirect('admin:student_subject_changelist')

        return render(
            request,
            'admin/subject_import.html',
            {
                'admin_site': self.admin_site,
                'title': 'Import Subjects from XLSX',
                'has_change_permission': self.has_change_permission(request),
            },
        )

    def download_xlsx_template(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Subjects'

        worksheet.append(['name', 'abbreviation', 'code', 'faculty'])
        worksheet.append(['Discrete Mathematics', 'DM', 'MATH101', 'Computer Science'])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="subject_import_template.xlsx"'
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        try:
            extra_context['import_xlsx_url'] = reverse(f'admin:{app_label}_{model_name}_import_xlsx')
            extra_context['download_template_url'] = reverse(f'admin:{app_label}_{model_name}_download_xlsx_template')
        except Exception:
            base = request.path.rsplit('/', 1)[0] + '/'
            extra_context['import_xlsx_url'] = base + 'import-xlsx/'
            extra_context['download_template_url'] = base + 'download-xlsx-template/'
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = ('name', 'roll_number', 'enrollment_batch', 'faculty', 'age', 'email', 'enrollment_date', 'user_column')
    search_fields = ('name', 'roll_number', '=enrollment_batch__year', 'faculty__name', 'email')
    list_filter = ('enrollment_batch', 'faculty', 'age', 'enrollment_date', 'is_first_login')
    change_list_template = 'admin/student/change_list.html'
    
    fieldsets = (
        ('Personal Information', {
            'fields': ('name', 'age', 'email', 'address', 'dp')
        }),
        ('Academic Information', {
            'fields': ('roll_number', 'enrollment_batch', 'faculty')
        }),
        ('Account Information', {
            'fields': ('user', 'is_first_login'),
            'classes': ('collapse',)
        }),
        ('Dates', {
            'fields': ('enrollment_date',),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('enrollment_date',)
    
    def user_column(self, obj):
        """Display user linked to student"""
        if obj.user:
            return format_html('<span style="color: green;">✓ {}</span>', obj.user.username)
        return mark_safe('<span style="color: red;">✗ No User</span>')
    user_column.short_description = 'User'
    
    def get_urls(self):
        """Add custom URLs for CSV import and download"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-csv/',
                self.admin_site.admin_view(self.import_csv),
                name='student_import_csv',
            ),
            path(
                'download-csv-template/',
                self.admin_site.admin_view(self.download_csv_template),
                name='student_download_csv_template',
            ),
        ]
        return custom_urls + urls
    
    def import_csv(self, request):
        """Handle CSV import request"""
        if request.method == 'POST':
            csv_file = request.FILES.get('csv_file')
            
            if not csv_file:
                messages.error(request, 'Please select a CSV file to import.')
                return render(request, 'admin/student_import.html', {'admin_site': self.admin_site})
            
            if not csv_file.name.lower().endswith('.csv'):
                messages.error(request, 'Only .csv files are supported.')
                return render(request, 'admin/student_import.html', {'admin_site': self.admin_site})
            
            # Import students from CSV
            created_users_data, updated_students_count, error_messages = import_students_from_csv(request, csv_file)
            
            # Display messages
            if error_messages:
                for error in error_messages:
                    messages.warning(request, error)
            
            if created_users_data:
                messages.success(
                    request,
                    f'Successfully created {len(created_users_data)} student account(s) '
                    f'and updated {updated_students_count} record(s).'
                )
                
                # Generate and download PDF with credentials
                pdf_buffer = generate_credentials_pdf(created_users_data)
                response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="student_credentials.pdf"'
                return response
            elif updated_students_count > 0:
                messages.success(request, f'Successfully updated {updated_students_count} student record(s).')
            else:
                messages.warning(request, 'No students were imported or updated.')
            
            return redirect('admin:student_student_changelist')
        
        # GET request - show import form
        return render(
            request,
            'admin/student_import.html',
            {
                'admin_site': self.admin_site,
                'title': 'Import Students from CSV',
                'has_change_permission': self.has_change_permission(request),
            }
        )
    
    def download_csv_template(self, request):
        """Download CSV template for student import"""
        return generate_csv_template(request)
    
    def changelist_view(self, request, extra_context=None):
        """Add import/download buttons to changelist"""
        extra_context = extra_context or {}
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        try:
            extra_context['import_csv_url'] = reverse(f'admin:{app_label}_{model_name}_import_csv')
            extra_context['download_template_url'] = reverse(f'admin:{app_label}_{model_name}_download_csv_template')
        except:
            # Fallback if reverse doesn't work
            base = request.path.rsplit('/', 1)[0] + '/'
            extra_context['import_csv_url'] = base + 'import-csv/'
            extra_context['download_template_url'] = base + 'download-csv-template/'
        return super().changelist_view(request, extra_context=extra_context)
