import csv
import uuid
from io import BytesIO, StringIO
from pathlib import Path
from collections import OrderedDict

from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from django.utils.decorators import method_decorator
from openpyxl import Workbook, load_workbook

from LMS.roles import get_student_profile, get_teacher_profile, is_admin_user
from LMS.upload_utils import validate_uploaded_file
from LMS.views import access_denied_response
from student.models import Faculty
from .models import Teacher
from .forms import FirstLoginPasswordChangeForm


class AdminOnlyMixin:
    def dispatch(self, request, *args, **kwargs):
        if not is_admin_user(request.user):
            return access_denied_response(request, 'Only administrators can access this page.')
        return super().dispatch(request, *args, **kwargs)


class TeacherListView(LoginRequiredMixin, ListView):
    model = Teacher
    template_name = 'teacher/teacher_list.html'
    context_object_name = 'teachers'
    paginate_by = 25

    def get_queryset(self):
        queryset = Teacher.objects.prefetch_related('faculties').order_by('name')
        search = (self.request.GET.get('q') or '').strip()
        faculty_id = (self.request.GET.get('faculty') or '').strip()

        if search:
            queryset = queryset.filter(name__icontains=search)
        if faculty_id:
            queryset = queryset.filter(faculties__id=faculty_id)

        if is_admin_user(self.request.user):
            return queryset
        if get_teacher_profile(self.request.user):
            return queryset
        return queryset.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        view_mode = self.request.GET.get('view', 'grouped')
        if view_mode not in {'grouped', 'flat'}:
            view_mode = 'grouped'

        context['view_mode'] = view_mode

        grouped_teachers_map = OrderedDict()
        for teacher in context['page_obj']:
            faculties = list(teacher.faculties.all())
            if not faculties:
                grouped_teachers_map.setdefault('No Faculty', []).append(teacher)
                continue

            for faculty in faculties:
                grouped_teachers_map.setdefault(faculty.name, []).append(teacher)

        context['grouped_teachers'] = [
            {
                'faculty': faculty,
                'teachers': teachers,
            }
            for faculty, teachers in grouped_teachers_map.items()
        ]
        query_dict = self.request.GET.copy()
        query_dict.pop('page', None)
        context['filter_query'] = query_dict.urlencode()
        context['search_query'] = (self.request.GET.get('q') or '').strip()
        context['selected_faculty'] = (self.request.GET.get('faculty') or '').strip()
        context['faculties'] = Faculty.objects.order_by('name')
        return context


class TeacherDetailView(LoginRequiredMixin, DetailView):
    model = Teacher
    template_name = 'teacher/teacher_detail.html'
    context_object_name = 'teacher'

    def get_queryset(self):
        queryset = Teacher.objects.prefetch_related('faculties', 'user')
        if is_admin_user(self.request.user) or get_teacher_profile(self.request.user):
            return queryset
        return queryset.none()


class TeacherCreateView(LoginRequiredMixin, AdminOnlyMixin, CreateView):
    model = Teacher
    template_name = 'teacher/teacher_form.html'
    fields = [
        'name',
        'employee_id',
        'department',
        'qualification',
        'experience_years',
        'faculties',
        'email',
        'dp',
        'address',
    ]
    success_url = reverse_lazy('teacher:teacher-create')


class TeacherUpdateView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):
    model = Teacher
    template_name = 'teacher/teacher_form.html'
    fields = [
        'name',
        'employee_id',
        'department',
        'qualification',
        'experience_years',
        'faculties',
        'email',
        'dp',
        'address',
    ]
    success_url = reverse_lazy('teacher:teacher-list')


class TeacherDeleteView(LoginRequiredMixin, AdminOnlyMixin, DeleteView):
    model = Teacher
    template_name = 'teacher/teacher_confirm_delete.html'
    success_url = reverse_lazy('teacher:teacher-list')


class TeacherImportView(LoginRequiredMixin, AdminOnlyMixin, View):
    template_name = 'teacher/teacher_import.html'
    IMPORT_SESSION_KEY = 'teacher_import_preview'

    def _stage_upload(self, upload):
        staging_dir = Path(settings.MEDIA_ROOT) / 'import_staging'
        staging_dir.mkdir(parents=True, exist_ok=True)
        file_name = f'teacher-{uuid.uuid4().hex}.xlsx'
        file_path = staging_dir / file_name
        with file_path.open('wb') as destination:
            for chunk in upload.chunks():
                destination.write(chunk)
        return str(file_path)

    def _cleanup_staged_file(self, file_path):
        if not file_path:
            return
        path_obj = Path(file_path)
        if path_obj.exists():
            path_obj.unlink()

    def _process_workbook(self, workbook, commit=False):
        worksheet = workbook.active
        created_count = 0
        updated_count = 0
        skipped_count = 0
        preview_rows = []
        error_rows = []

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value in (None, '') for value in row):
                continue

            name = str(row[0]).strip() if row[0] else ''
            employee_id = str(row[1]).strip() if row[1] else None
            department = str(row[2]).strip() if row[2] else None
            qualification = str(row[3]).strip() if row[3] else None
            experience_raw = row[4]
            faculties_raw = str(row[5]).strip() if row[5] else ''
            email = str(row[6]).strip() if row[6] else None
            address = str(row[7]).strip() if row[7] else None

            if not name:
                skipped_count += 1
                error_rows.append({'row': row_index, 'error': 'Missing required teacher name.'})
                continue

            experience_years = None
            if experience_raw not in (None, ''):
                try:
                    experience_years = int(experience_raw)
                except (TypeError, ValueError):
                    skipped_count += 1
                    error_rows.append({'row': row_index, 'error': 'Invalid experience year value.'})
                    continue

            defaults = {
                'name': name,
                'department': department,
                'qualification': qualification,
                'experience_years': experience_years,
                'email': email,
                'address': address,
            }

            operation = 'create'
            if employee_id and Teacher.objects.filter(employee_id=employee_id).exists():
                operation = 'update'
            elif not employee_id and email and Teacher.objects.filter(email=email).exists():
                operation = 'update'

            if len(preview_rows) < 100:
                preview_rows.append(
                    {
                        'row': row_index,
                        'name': name,
                        'employee_id': employee_id or '-',
                        'email': email or '-',
                        'operation': operation,
                    }
                )

            if not commit:
                if operation == 'create':
                    created_count += 1
                else:
                    updated_count += 1
                continue

            try:
                if employee_id:
                    teacher, created = Teacher.objects.update_or_create(
                        employee_id=employee_id,
                        defaults=defaults,
                    )
                elif email:
                    teacher, created = Teacher.objects.update_or_create(
                        email=email,
                        defaults={**defaults, 'employee_id': None},
                    )
                else:
                    teacher = Teacher.objects.create(
                        **defaults,
                        employee_id=None,
                    )
                    created = True
            except Exception:
                skipped_count += 1
                error_rows.append({'row': row_index, 'error': 'Database write failed for this row.'})
                continue

            faculty_names = [item.strip() for item in faculties_raw.split(',') if item.strip()]
            faculties = [
                Faculty.objects.get_or_create_case_insensitive(faculty_name)[0]
                for faculty_name in faculty_names
            ]
            teacher.faculties.set(faculties)

            if created:
                created_count += 1
            else:
                updated_count += 1

        return {
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'preview_rows': preview_rows,
            'error_rows': error_rows,
        }

    def _get_preview_context(self, request):
        return request.session.get(self.IMPORT_SESSION_KEY, {})

    def get(self, request):
        context = self._get_preview_context(request)
        return render(request, self.template_name, context)

    def post(self, request):
        action = request.POST.get('action', 'preview')

        if action == 'commit':
            preview_data = request.session.get(self.IMPORT_SESSION_KEY)
            if not preview_data or not preview_data.get('staged_file_path'):
                messages.error(request, 'No staged import found. Please preview an XLSX file first.')
                return redirect('teacher:teacher-import')

            try:
                workbook = load_workbook(preview_data['staged_file_path'], data_only=True)
            except Exception:
                self._cleanup_staged_file(preview_data.get('staged_file_path'))
                request.session.pop(self.IMPORT_SESSION_KEY, None)
                messages.error(request, 'Unable to read staged file. Please upload and preview again.')
                return redirect('teacher:teacher-import')

            result = self._process_workbook(workbook, commit=True)
            self._cleanup_staged_file(preview_data.get('staged_file_path'))
            request.session.pop(self.IMPORT_SESSION_KEY, None)

            messages.success(
                request,
                f'Import committed. Created: {result["created_count"]}, Updated: {result["updated_count"]}, Skipped: {result["skipped_count"]}.',
            )
            return redirect('teacher:teacher-list')

        upload = request.FILES.get('xlsx_file')
        try:
            validate_uploaded_file(
                upload,
                allowed_extensions={'.xlsx'},
                allowed_mime_types={
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/octet-stream',
                },
                max_size_bytes=settings.MAX_IMPORT_UPLOAD_BYTES,
            )
        except Exception as exc:
            messages.error(request, str(exc))
            return redirect('teacher:teacher-import')

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            messages.error(request, 'Unable to read the file. Please use the provided template.')
            return redirect('teacher:teacher-import')

        previous_staged_file = request.session.get(self.IMPORT_SESSION_KEY, {}).get('staged_file_path')
        self._cleanup_staged_file(previous_staged_file)
        staged_file_path = self._stage_upload(upload)
        result = self._process_workbook(workbook, commit=False)
        request.session[self.IMPORT_SESSION_KEY] = {
            'staged_file_path': staged_file_path,
            'uploaded_file_name': upload.name,
            'created_count': result['created_count'],
            'updated_count': result['updated_count'],
            'skipped_count': result['skipped_count'],
            'preview_rows': result['preview_rows'],
            'error_rows': result['error_rows'],
        }

        messages.info(request, 'Preview generated. Review changes and commit when ready.')
        return redirect('teacher:teacher-import')


class TeacherImportErrorReportView(LoginRequiredMixin, AdminOnlyMixin, View):
    def get(self, request):
        preview_data = request.session.get(TeacherImportView.IMPORT_SESSION_KEY)
        if not preview_data:
            messages.error(request, 'No preview is available for download.')
            return redirect('teacher:teacher-import')

        error_rows = preview_data.get('error_rows', [])
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['row', 'error'])
        for error in error_rows:
            writer.writerow([error.get('row', ''), error.get('error', '')])

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="teacher_import_errors.csv"'
        return response


class TeacherImportTemplateView(LoginRequiredMixin, AdminOnlyMixin, View):
    def get(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Teachers'

        worksheet.append([
            'name',
            'employee_id',
            'department',
            'qualification',
            'experience_years',
            'faculties',
            'email',
            'address',
        ])
        worksheet.append([
            'Sita Karki',
            'EMP-1042',
            'Computer Science',
            'MSc Computer Science',
            6,
            'Computer Science, IT',
            'sita.karki@example.com',
            'Pokhara',
        ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="teacher_import_template.xlsx"'
        return response


@method_decorator(login_required, name='dispatch')
class FirstLoginPasswordChangeView(View):
    """View for changing password on first login"""
    template_name = 'teacher/first_login_password_change.html'
    
    def get(self, request):
        """Display password change form"""
        try:
            teacher = request.user.teacher_profile
            if not teacher.is_first_login:
                messages.info(request, 'You have already changed your password.')
                return redirect('teacher:teacher-list')
        except AttributeError:
            # User is not a teacher, redirect to home
            messages.error(request, 'Only teachers can access this page.')
            return redirect('/')
        
        form = FirstLoginPasswordChangeForm(request.user)
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        """Process password change"""
        try:
            teacher = request.user.teacher_profile
        except AttributeError:
            messages.error(request, 'Only teachers can access this page.')
            return redirect('/')
        
        form = FirstLoginPasswordChangeForm(request.user, request.POST)
        
        if form.is_valid():
            form.save()
            
            # Mark first login as False
            teacher.is_first_login = False
            teacher.save()
            
            messages.success(request, 'Your password has been changed successfully. Please login again.')
            return redirect('login')
        
        return render(request, self.template_name, {'form': form})

