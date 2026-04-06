from io import BytesIO
from datetime import timedelta
from collections import OrderedDict

from django.contrib import messages
from django import forms
from django.db.models import Q
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from student.models import EnrollmentYear, Faculty, Student, Subject
from LMS.roles import is_admin_user
from LMS.views import access_denied_response

from .models import Attendance, AttendanceEntry, TimeSlot, WeeklyClassSchedule


WEEKDAY_INDEX = {
    'monday': 0,
    'tuesday': 1,
    'wednesday': 2,
    'thursday': 3,
    'friday': 4,
    'saturday': 5,
    'sunday': 6,
}


def get_accessible_schedule_queryset(user):
    schedules = WeeklyClassSchedule.objects.filter(is_active=True).select_related(
        'subject', 'teacher', 'faculty', 'enrollment_batch', 'timeslot'
    )

    if is_admin_user(user):
        return schedules

    student_profile = getattr(user, 'student_profile', None)
    if student_profile:
        return schedules.filter(
            faculty=student_profile.faculty,
            enrollment_batch=student_profile.enrollment_batch,
        )

    teacher_profile = getattr(user, 'teacher_profile', None)
    if teacher_profile:
        return schedules.filter(
            Q(teacher=teacher_profile) | Q(faculty__in=teacher_profile.faculties.all())
        ).distinct()

    return schedules.none()


def sync_attendance_from_schedules(user, until_date):
    schedules = get_accessible_schedule_queryset(user)

    for schedule in schedules:
        if not all([schedule.subject, schedule.faculty, schedule.enrollment_batch, schedule.timeslot, schedule.start_date]):
            continue

        day_index = WEEKDAY_INDEX.get(schedule.day_of_week)
        if day_index is None:
            continue

        range_end = min(until_date, schedule.end_date) if schedule.end_date else until_date
        if schedule.start_date > range_end:
            continue

        days_ahead = (day_index - schedule.start_date.weekday()) % 7
        current_date = schedule.start_date + timedelta(days=days_ahead)

        while current_date <= range_end:
            attendance, _ = Attendance.objects.get_or_create(
                subject=schedule.subject,
                teacher=schedule.teacher,
                faculty=schedule.faculty,
                enrollment_batch=schedule.enrollment_batch,
                timeslot=schedule.timeslot,
                attendance_date=current_date,
                defaults={
                    'note': schedule.note,
                },
            )
            if not attendance.note and schedule.note:
                attendance.note = schedule.note
                attendance.save(update_fields=['note'])
            current_date += timedelta(days=7)


def seed_attendance_entries(attendance):
    if not attendance.faculty or not attendance.enrollment_batch:
        return

    students = Student.objects.filter(
        faculty=attendance.faculty,
        enrollment_batch=attendance.enrollment_batch,
    )
    for student in students:
        AttendanceEntry.objects.get_or_create(
            attendance=attendance,
            student=student,
        )


class AttendanceForm(forms.ModelForm):
    class Meta:
        model = Attendance
        fields = ['subject', 'teacher', 'faculty', 'enrollment_batch', 'attendance_date', 'timeslot', 'note']
        widgets = {
            'attendance_date': forms.DateInput(attrs={'type': 'date'}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['timeslot'].queryset = TimeSlot.objects.order_by('display_order', 'start_time', 'id')

    def clean(self):
        cleaned_data = super().clean()
        subject = cleaned_data.get('subject')
        faculty = cleaned_data.get('faculty')
        enrollment_batch = cleaned_data.get('enrollment_batch')
        attendance_date = cleaned_data.get('attendance_date')
        timeslot = cleaned_data.get('timeslot')

        if attendance_date and attendance_date > timezone.localdate():
            raise forms.ValidationError('Attendance for future dates cannot be created yet.')

        if subject and faculty and enrollment_batch and attendance_date and timeslot:
            duplicate_exists = Attendance.objects.filter(
                subject=subject,
                faculty=faculty,
                enrollment_batch=enrollment_batch,
                attendance_date=attendance_date,
                timeslot=timeslot,
            )
            if self.instance.pk:
                duplicate_exists = duplicate_exists.exclude(pk=self.instance.pk)

            if duplicate_exists.exists():
                raise forms.ValidationError(
                    'An attendance record for this subject, faculty, year, date, and timeslot already exists.'
                )

        return cleaned_data


class AttendanceReportForm(forms.Form):
    subject = forms.ModelChoiceField(queryset=Subject.objects.order_by('name'))
    start_date = forms.DateField(widget=forms.DateInput(attrs={'type': 'date'}))
    end_date = forms.DateField(widget=forms.DateInput(attrs={'type': 'date'}))


class RoleFilteredAttendanceQuerysetMixin:
    def get_base_queryset(self):
        return Attendance.objects.select_related(
            'subject', 'teacher', 'faculty', 'enrollment_batch', 'timeslot'
        ).prefetch_related('entries__student')

    def get_role_filtered_queryset(self):
        sync_attendance_from_schedules(self.request.user, timezone.localdate())
        queryset = self.get_base_queryset()
        user = self.request.user

        queryset = queryset.filter(attendance_date__lte=timezone.localdate())

        if user.is_superuser or user.is_staff:
            return queryset

        student_profile = getattr(user, 'student_profile', None)
        if student_profile:
            return queryset.filter(entries__student=student_profile).distinct()

        teacher_profile = getattr(user, 'teacher_profile', None)
        if teacher_profile:
            return queryset.filter(
                Q(teacher=teacher_profile) | Q(faculty__in=teacher_profile.faculties.all())
            ).distinct()

        return queryset.none()


class TeacherOrAdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        user = self.request.user
        return user.is_authenticated and (
            user.is_superuser or user.is_staff or hasattr(user, 'teacher_profile')
        )

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            return access_denied_response(self.request, 'You do not have permission to modify attendance records.')
        return super().handle_no_permission()


class AttendanceListView(LoginRequiredMixin, RoleFilteredAttendanceQuerysetMixin, ListView):
    model = Attendance
    template_name = 'attendance/attendance_list.html'
    context_object_name = 'attendances'
    paginate_by = 25

    def get_queryset(self):
        queryset = self.get_role_filtered_queryset().order_by('-attendance_date', '-created_at')
        faculty_id = (self.request.GET.get('faculty') or '').strip()
        year_id = (self.request.GET.get('year') or '').strip()
        subject_id = (self.request.GET.get('subject') or '').strip()

        if faculty_id:
            queryset = queryset.filter(faculty_id=faculty_id)
        if year_id:
            queryset = queryset.filter(enrollment_batch_id=year_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query_dict = self.request.GET.copy()
        query_dict.pop('page', None)
        context['filter_query'] = query_dict.urlencode()
        context['selected_faculty'] = (self.request.GET.get('faculty') or '').strip()
        context['selected_year'] = (self.request.GET.get('year') or '').strip()
        context['selected_subject'] = (self.request.GET.get('subject') or '').strip()
        context['faculties'] = Faculty.objects.order_by('name')
        context['enrollment_years'] = EnrollmentYear.objects.order_by('year')
        context['subjects'] = Subject.objects.order_by('name')
        return context


class AttendanceDetailView(LoginRequiredMixin, RoleFilteredAttendanceQuerysetMixin, DetailView):
    model = Attendance
    template_name = 'attendance/attendance_detail.html'
    context_object_name = 'attendance'

    def get_queryset(self):
        return self.get_role_filtered_queryset()


class AttendanceRosterView(LoginRequiredMixin, RoleFilteredAttendanceQuerysetMixin, DetailView):
    model = Attendance
    template_name = 'attendance/attendance_roster.html'
    context_object_name = 'attendance'

    def get_queryset(self):
        queryset = self.get_role_filtered_queryset().select_related('subject', 'teacher', 'faculty', 'enrollment_batch')
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return queryset

        teacher_profile = getattr(user, 'teacher_profile', None)
        if teacher_profile:
            return queryset.filter(
                Q(teacher=teacher_profile) | Q(faculty__in=teacher_profile.faculties.all())
            ).distinct()

        return queryset.none()

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not (request.user.is_superuser or request.user.is_staff or hasattr(request.user, 'teacher_profile')):
            return access_denied_response(request, 'Only teachers and admins can view attendance rosters.')
        if request.method == 'POST' and self.object.attendance_date > timezone.localdate():
            return access_denied_response(request, 'Attendance can only be edited on its date or after the date arrives.')
        self.ensure_entries_exist()
        if request.method == 'POST':
            return self.handle_post(request)
        return super().dispatch(request, *args, **kwargs)

    def ensure_entries_exist(self):
        seed_attendance_entries(self.object)

    def handle_post(self, request):
        updated_count = 0
        for entry in self.object.entries.all():
            status = request.POST.get(f'status_{entry.pk}')
            if status not in {'present', 'absent', 'unmarked'}:
                continue

            new_marked_at = timezone.now() if status != 'unmarked' else None
            if entry.status != status or entry.marked_at != new_marked_at:
                entry.status = status
                entry.marked_at = new_marked_at
                entry.save(update_fields=['status', 'marked_at'])
                updated_count += 1

        if updated_count:
            messages.success(
                self.request,
                f"Attendance saved for {updated_count} student{'s' if updated_count != 1 else ''}.",
            )
        else:
            messages.info(self.request, 'No attendance changes to save.')
        return redirect('attendance:attendance-roster', pk=self.object.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['entries'] = self.object.entries.select_related('student').order_by('student__name')
        context['matching_students'] = Student.objects.filter(
            faculty=self.object.faculty,
            enrollment_batch=self.object.enrollment_batch,
        ).order_by('name') if self.object.faculty and self.object.enrollment_batch else Student.objects.none()
        return context


class AttendanceCreateView(LoginRequiredMixin, TeacherOrAdminRequiredMixin, CreateView):
    model = Attendance
    template_name = 'attendance/attendance_form.html'
    form_class = AttendanceForm

    def form_valid(self, form):
        self.object = form.save()
        seed_attendance_entries(self.object)
        messages.success(self.request, 'Attendance created successfully.')
        return redirect('attendance:attendance-roster', pk=self.object.pk)


class AttendanceUpdateView(LoginRequiredMixin, TeacherOrAdminRequiredMixin, RoleFilteredAttendanceQuerysetMixin, UpdateView):
    model = Attendance
    template_name = 'attendance/attendance_form.html'
    form_class = AttendanceForm
    success_url = reverse_lazy('attendance:attendance-list')

    def dispatch(self, request, *args, **kwargs):
        attendance = self.get_object()
        if attendance.attendance_date > timezone.localdate():
            return access_denied_response(request, 'Future attendance records cannot be edited yet.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = self.get_role_filtered_queryset()
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return queryset
        teacher_profile = getattr(user, 'teacher_profile', None)
        if teacher_profile:
            return queryset.filter(
                Q(teacher=teacher_profile) | Q(faculty__in=teacher_profile.faculties.all())
            ).distinct()
        return queryset.none()


class AttendanceDeleteView(LoginRequiredMixin, TeacherOrAdminRequiredMixin, RoleFilteredAttendanceQuerysetMixin, DeleteView):
    model = Attendance
    template_name = 'attendance/attendance_confirm_delete.html'
    success_url = reverse_lazy('attendance:attendance-list')

    def dispatch(self, request, *args, **kwargs):
        attendance = self.get_object()
        if attendance.attendance_date > timezone.localdate():
            return access_denied_response(request, 'Future attendance records cannot be deleted yet.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = self.get_role_filtered_queryset()
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return queryset
        teacher_profile = getattr(user, 'teacher_profile', None)
        if teacher_profile:
            return queryset.filter(
                Q(teacher=teacher_profile) | Q(faculty__in=teacher_profile.faculties.all())
            ).distinct()
        return queryset.none()


class AttendanceReportView(LoginRequiredMixin, RoleFilteredAttendanceQuerysetMixin, View):
    template_name = 'attendance/attendance_report.html'

    def get(self, request):
        form = AttendanceReportForm(request.GET or None)
        form.fields['subject'].queryset = self.get_subject_queryset()
        return render(request, self.template_name, {'form': form})

    def get_subject_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return Subject.objects.order_by('name')

        attendance_queryset = self.get_role_filtered_queryset()
        subject_ids = attendance_queryset.values_list('subject_id', flat=True)
        return Subject.objects.filter(id__in=subject_ids).order_by('name').distinct()


class AttendanceReportDownloadView(LoginRequiredMixin, RoleFilteredAttendanceQuerysetMixin, View):
    MAX_REPORT_RANGE_DAYS = 62

    def post(self, request):
        form = AttendanceReportForm(request.POST)
        form.fields['subject'].queryset = self._allowed_subject_queryset()
        if not form.is_valid():
            messages.error(request, 'Please choose a valid subject and date range.')
            return redirect('attendance:attendance-report')

        subject = form.cleaned_data['subject']
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        report_format = request.POST.get('format', 'xlsx')

        if start_date > end_date:
            messages.error(request, 'Start date cannot be after end date.')
            return redirect('attendance:attendance-report')

        day_span = (end_date - start_date).days + 1
        if day_span > self.MAX_REPORT_RANGE_DAYS:
            messages.error(
                request,
                f'Date range is too large ({day_span} days). Please use a range up to {self.MAX_REPORT_RANGE_DAYS} days.',
            )
            return redirect('attendance:attendance-report')

        allowed_attendance_queryset = self._allowed_attendance_queryset(subject, start_date, end_date)
        report_dates = self._build_date_list(start_date, end_date)
        report_rows = self._build_matrix(allowed_attendance_queryset, report_dates)
        title_text = f'{subject.name} Attendance Report {start_date} - {end_date}'
        filename_base = f'attendance_report_{subject.name}_{start_date}_to_{end_date}'.replace(' ', '_')

        if report_format == 'pdf':
            return self._build_pdf_response(filename_base, title_text, report_dates, report_rows)
        return self._build_xlsx_response(filename_base, title_text, report_dates, report_rows)

    def _build_date_list(self, start_date, end_date):
        report_dates = []
        current_date = start_date
        while current_date <= end_date:
            report_dates.append(current_date)
            current_date += timedelta(days=1)
        return report_dates

    def _allowed_subject_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return Subject.objects.order_by('name')

        attendance_queryset = self.get_role_filtered_queryset()
        subject_ids = attendance_queryset.values_list('subject_id', flat=True)
        return Subject.objects.filter(id__in=subject_ids).order_by('name').distinct()

    def _allowed_attendance_queryset(self, subject, start_date, end_date):
        return self.get_role_filtered_queryset().filter(
            subject=subject,
            attendance_date__range=(start_date, end_date),
        ).order_by('attendance_date', 'id')

    def _build_matrix(self, attendances, report_dates):
        user = self.request.user
        student_profile = getattr(user, 'student_profile', None)

        attendance_lookup = {}
        status_lookup = {}
        students_by_id = {}

        for attendance in attendances:
            attendance_lookup[(attendance.attendance_date, attendance.faculty_id, attendance.enrollment_batch_id)] = attendance
            for entry in attendance.entries.all():
                if student_profile and entry.student_id != student_profile.id:
                    continue
                students_by_id[entry.student_id] = entry.student
                status_lookup[(attendance.attendance_date, attendance.faculty_id, attendance.enrollment_batch_id, entry.student_id)] = entry.status

        students = sorted(
            students_by_id.values(),
            key=lambda student: (
                student.roll_number or '',
                student.name or '',
            ),
        )

        rows = []
        total_days = len(report_dates)
        for student in students:
            statuses = []
            present_count = 0
            for report_date in report_dates:
                attendance = attendance_lookup.get((report_date, student.faculty_id, student.enrollment_batch_id))
                status = '-'
                if attendance:
                    status = status_lookup.get((attendance.attendance_date, attendance.faculty_id, attendance.enrollment_batch_id, student.id), '-')

                if status == AttendanceEntry.STATUS_PRESENT:
                    statuses.append('P')
                    present_count += 1
                elif status == AttendanceEntry.STATUS_ABSENT:
                    statuses.append('A')
                else:
                    statuses.append('-')

            percentage_ratio = (present_count / total_days) if total_days else 0
            rows.append({
                'roll_no': student.roll_number or '-',
                'name': student.name,
                'statuses': statuses,
                'present_count': present_count,
                'percentage_ratio': percentage_ratio,
                'percentage_display': f'{percentage_ratio * 100:.2f}%',
                'highlight': percentage_ratio < 0.5,
            })

        return rows

    def _build_xlsx_response(self, filename_base, title_text, report_dates, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Attendance Report'

        total_columns = 3 + len(report_dates)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        title_cell = worksheet.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = PatternFill('solid', fgColor='EAF2FF')

        date_range_cell = worksheet.cell(
            row=2,
            column=1,
            value=f'Date Range: {report_dates[0]} to {report_dates[-1]}' if report_dates else 'Date Range: -',
        )
        date_range_cell.font = Font(italic=True)

        worksheet.append([])
        headers = ['Roll No', 'Name'] + [report_date.strftime('%d-%b') for report_date in report_dates] + ['Percentage']
        worksheet.append(headers)

        for cell in worksheet[4]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='155EEF')

        low_fill = PatternFill('solid', fgColor='FDE2E1')

        for row_index, row in enumerate(rows, start=5):
            row_values = [row['roll_no'], row['name'], *row['statuses'], row['percentage_ratio']]
            worksheet.append(row_values)

            if row['highlight']:
                for column_index in range(1, total_columns + 1):
                    worksheet.cell(row=row_index, column=column_index).fill = low_fill

            percentage_cell = worksheet.cell(row=row_index, column=total_columns)
            percentage_cell.number_format = '0.00%'
            percentage_cell.alignment = percentage_cell.alignment.copy(horizontal='center')

        worksheet.freeze_panes = 'C5'
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 50
        for index in range(3, total_columns):
            worksheet.column_dimensions[get_column_letter(index)].width = 10
        worksheet.column_dimensions[get_column_letter(total_columns)].width = 12

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    def _build_pdf_response(self, filename_base, title_text, report_dates, rows):
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )
        styles = getSampleStyleSheet()

        story = [
            Paragraph(title_text, styles['Title']),
            Spacer(1, 12),
            Paragraph(
                f'Date Range: {report_dates[0]} to {report_dates[-1]}' if report_dates else 'Date Range: -',
                styles['Normal'],
            ),
            Spacer(1, 12),
        ]

        headers = ['Roll No', 'Name'] + [report_date.strftime('%d-%b') for report_date in report_dates] + ['Percentage']
        table_data = [headers]
        if rows:
            for row in rows:
                table_data.append([
                    row['roll_no'],
                    row['name'],
                    *row['statuses'],
                    row['percentage_display'],
                ])
        else:
            table_data.append(['No matching attendance records found.'] + [''] * (len(headers) - 1))

        page_width = landscape(A4)[0] - document.leftMargin - document.rightMargin
        fixed_widths = [90, 220, 58]
        remaining_width = max(page_width - sum(fixed_widths), 0)
        date_width = (remaining_width / len(report_dates)) if report_dates else 0
        column_widths = [fixed_widths[0], fixed_widths[1]] + [date_width for _ in report_dates] + [fixed_widths[2]]

        table = Table(table_data, repeatRows=1, colWidths=column_widths)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#155eef')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ])

        for row_index, row in enumerate(rows, start=1):
            if row['highlight']:
                table_style.add('BACKGROUND', (0, row_index), (-1, row_index), colors.HexColor('#FDE2E1'))
                table_style.add('TEXTCOLOR', (0, row_index), (-1, row_index), colors.HexColor('#7A1E1E'))

        table.setStyle(table_style)
        story.append(table)

        document.build(story)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
        return response


class AdminOnlyMixin(UserPassesTestMixin):
    def test_func(self):
        return is_admin_user(self.request.user)

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            return access_denied_response(self.request, 'Only administrators can access this module.')
        return super().handle_no_permission()


class WeeklyScheduleForm(forms.ModelForm):
    class Meta:
        model = WeeklyClassSchedule
        fields = [
            'subject',
            'teacher',
            'faculty',
            'enrollment_batch',
            'day_of_week',
            'timeslot',
            'start_date',
            'end_date',
            'note',
            'is_active',
        ]
        widgets = {
            'start_date': forms.DateInput(attrs={'type': 'date'}),
            'end_date': forms.DateInput(attrs={'type': 'date'}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['timeslot'].queryset = TimeSlot.objects.order_by('display_order', 'start_time', 'id')


class WeeklyScheduleListView(LoginRequiredMixin, AdminOnlyMixin, ListView):
    model = WeeklyClassSchedule
    template_name = 'attendance/weekly_schedule_list.html'
    context_object_name = 'schedules'

    def get_queryset(self):
        return WeeklyClassSchedule.objects.select_related(
            'subject', 'teacher', 'faculty', 'enrollment_batch', 'timeslot'
        ).order_by('faculty__name', 'enrollment_batch__year', 'day_of_week', 'timeslot__display_order', 'timeslot__start_time')


class WeeklyScheduleCreateView(LoginRequiredMixin, AdminOnlyMixin, CreateView):
    model = WeeklyClassSchedule
    template_name = 'attendance/weekly_schedule_form.html'
    form_class = WeeklyScheduleForm

    def get_success_url(self):
        # If coming from timetable, go back to timetable
        if self.request.GET.get('day'):
            faculty_id = self.request.GET.get('faculty_id', '')
            year_id = self.request.GET.get('year_id', '')
            return f"{reverse_lazy('attendance:attendance-timetable')}?faculty={faculty_id}&year={year_id}"
        return reverse_lazy('attendance:weekly-schedule-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        faculty_id = self.request.GET.get('faculty_id')
        year_id = self.request.GET.get('year_id')
        day = self.request.GET.get('day')
        timeslot_id = self.request.GET.get('timeslot_id')

        if faculty_id:
            initial['faculty'] = faculty_id
        if year_id:
            initial['enrollment_batch'] = year_id
        if day:
            initial['day_of_week'] = day
        if timeslot_id:
            initial['timeslot'] = timeslot_id

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass query params to template for back button
        context['faculty_id'] = self.request.GET.get('faculty_id', '')
        context['year_id'] = self.request.GET.get('year_id', '')
        context['day'] = self.request.GET.get('day', '')
        context['timeslot_id'] = self.request.GET.get('timeslot_id', '')
        context['return_to_timetable'] = bool(self.request.GET.get('day'))
        return context


class WeeklyScheduleUpdateView(LoginRequiredMixin, AdminOnlyMixin, UpdateView):
    model = WeeklyClassSchedule
    template_name = 'attendance/weekly_schedule_form.html'
    form_class = WeeklyScheduleForm

    def get_success_url(self):
        # If coming from timetable, go back to timetable
        if self.request.GET.get('day'):
            faculty_id = self.request.GET.get('faculty_id', '')
            year_id = self.request.GET.get('year_id', '')
            return f"{reverse_lazy('attendance:attendance-timetable')}?faculty={faculty_id}&year={year_id}"
        return reverse_lazy('attendance:weekly-schedule-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass query params to template for back button
        context['faculty_id'] = self.request.GET.get('faculty_id', '')
        context['year_id'] = self.request.GET.get('year_id', '')
        context['day'] = self.request.GET.get('day', '')
        context['timeslot_id'] = self.request.GET.get('timeslot_id', '')
        context['return_to_timetable'] = bool(self.request.GET.get('day'))
        return context


class WeeklyScheduleDeleteView(LoginRequiredMixin, AdminOnlyMixin, DeleteView):
    model = WeeklyClassSchedule
    template_name = 'attendance/weekly_schedule_confirm_delete.html'
    success_url = reverse_lazy('attendance:weekly-schedule-list')


class AttendanceTimetableView(LoginRequiredMixin, ListView):
    model = WeeklyClassSchedule
    template_name = 'attendance/attendance_timetable.html'
    context_object_name = 'schedules'
    WEEKDAY_ORDER = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']

    def _filter_options(self):
        user = self.request.user
        if is_admin_user(user):
            faculties = Faculty.objects.order_by('name')
            years = EnrollmentYear.objects.order_by('year')
        else:
            teacher_profile = getattr(user, 'teacher_profile', None)
            student_profile = getattr(user, 'student_profile', None)
            if teacher_profile:
                faculties = teacher_profile.faculties.order_by('name')
                years = EnrollmentYear.objects.filter(students__faculty__in=faculties).distinct().order_by('year')
            elif student_profile:
                faculties = Faculty.objects.filter(pk=student_profile.faculty_id)
                years = EnrollmentYear.objects.filter(pk=student_profile.enrollment_batch_id)
            else:
                faculties = Faculty.objects.none()
                years = EnrollmentYear.objects.none()
        return faculties, years

    def get_queryset(self):
        queryset = get_accessible_schedule_queryset(self.request.user)
        user = self.request.user
        faculty_id = self.request.GET.get('faculty')
        year_id = self.request.GET.get('year')

        # For students, auto-apply their batch and faculty filters if not manually overridden
        student_profile = getattr(user, 'student_profile', None)
        if student_profile:
            if not faculty_id:
                faculty_id = str(student_profile.faculty_id)
            if not year_id:
                year_id = str(student_profile.enrollment_batch_id)

        if faculty_id:
            queryset = queryset.filter(faculty_id=faculty_id)
        if year_id:
            queryset = queryset.filter(enrollment_batch_id=year_id)

        return queryset.order_by('day_of_week', 'timeslot__display_order', 'timeslot__start_time', 'subject__name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ordered_slots = list(TimeSlot.objects.order_by('display_order', 'start_time', 'id'))
        context['timeslots'] = ordered_slots

        faculties, years = self._filter_options()
        context['faculties'] = faculties
        context['enrollment_years'] = years
        
        # For students, pre-select their faculty and batch
        selected_faculty = self.request.GET.get('faculty', '')
        selected_year = self.request.GET.get('year', '')
        
        user = self.request.user
        student_profile = getattr(user, 'student_profile', None)
        if student_profile and not selected_faculty and not selected_year:
            selected_faculty = str(student_profile.faculty_id)
            selected_year = str(student_profile.enrollment_batch_id)
        
        context['selected_faculty'] = selected_faculty
        context['selected_year'] = selected_year
        
        # Check if both filters are selected
        has_both_filters = bool(selected_faculty and selected_year)
        context['has_both_filters'] = has_both_filters

        # Only build timetable if both filters are selected
        if has_both_filters:
            schedule_index = {}
            for schedule in context['schedules']:
                if schedule.day_of_week and schedule.timeslot_id:
                    schedule_index[(schedule.day_of_week, schedule.timeslot_id)] = schedule

            rows = []
            for day_key in self.WEEKDAY_ORDER:
                cells = []
                for slot in ordered_slots:
                    schedule = schedule_index.get((day_key, slot.id))
                    cells.append({
                        'slot': slot,
                        'schedule': schedule,
                        'is_break': slot.is_break,
                    })
                rows.append({
                    'day_key': day_key,
                    'day_label': day_key.title(),
                    'cells': cells,
                })

            context['timetable_rows'] = rows
        else:
            context['timetable_rows'] = []
        
        return context

    def _build_pdf_response(self, filename_base, title_text, report_dates, rows):
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )
        styles = getSampleStyleSheet()

        story = [
            Paragraph(title_text, styles['Title']),
            Spacer(1, 12),
            Paragraph(
                f'Date Range: {report_dates[0]} to {report_dates[-1]}' if report_dates else 'Date Range: -',
                styles['Normal'],
            ),
            Spacer(1, 12),
        ]

        headers = ['Roll No', 'Name'] + [report_date.strftime('%d-%b') for report_date in report_dates] + ['Percentage']
        table_data = [headers]
        if rows:
            for row in rows:
                table_data.append([
                    row['roll_no'],
                    row['name'],
                    *row['statuses'],
                    row['percentage_display'],
                ])
        else:
            table_data.append(['No matching attendance records found.'] + [''] * (len(headers) - 1))

        page_width = landscape(A4)[0] - document.leftMargin - document.rightMargin
        fixed_widths = [90, 220, 58]
        remaining_width = max(page_width - sum(fixed_widths), 0)
        date_width = (remaining_width / len(report_dates)) if report_dates else 0
        column_widths = [fixed_widths[0], fixed_widths[1]] + [date_width for _ in report_dates] + [fixed_widths[2]]

        table = Table(table_data, repeatRows=1, colWidths=column_widths)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#155eef')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ])

        for row_index, row in enumerate(rows, start=1):
            if row['highlight']:
                table_style.add('BACKGROUND', (0, row_index), (-1, row_index), colors.HexColor('#FDE2E1'))
                table_style.add('TEXTCOLOR', (0, row_index), (-1, row_index), colors.HexColor('#7A1E1E'))

        table.setStyle(table_style)
        story.append(table)

        document.build(story)
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
        return response
